from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .measurement import SensorStore

COLUMNS = [
    "Timestamp", "Plant", "Building/Floor/Section", "Asset Ref.",
    "Sensor", "Sensor Status", "Test Duration",
    "RSSI (dBm)", "RSRP (dBm)", "RSRQ (dB)", "SINR (dB)", "Speed (kB/s)",
    "MCC/MNC", "LAC", "CellID", "Access Technology",
    "Band", "Channel Type", "SM /2G", "Signal Quality",
]
N_COLS = len(COLUMNS)

_HDR_FILL = PatternFill("solid", fgColor="1A2B4A")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_CENTER   = Alignment(horizontal="center", vertical="center")
_THIN     = Side(style="thin", color="D1D9E6")
_BORDER   = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)


def _blank(v):
    return v if v is not None else ""


def safe_name(s: str) -> str:
    return re.sub(r"_+", "_",
        "".join("_" if c in '<>:"/\\|?*\n\r\t' else c for c in s).strip()) or "x"


def default_filename(plant: str, building: str) -> str:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"{safe_name(plant) or 'Plant'}_{safe_name(building) or 'Building'}_{ts}.xlsx"


def _parse_elapsed(mmss: str) -> int:
    try:
        m, s = mmss.split(":")
        return int(m) * 60 + int(s)
    except Exception:
        return 0


def _write_sheet(ws, store: SensorStore, sensor_label: str,
                 plant: str, building: str, asset: str) -> None:
    # Header row
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    # Build timeline (samples + events, sorted by elapsed time)
    # Placed first so the data is easy to read top-to-bottom.
    timeline = []
    for s in store.samples:
        timeline.append((_parse_elapsed(s.elapsed), 1, "sample", s))
    for e in store.events:
        timeline.append((_parse_elapsed(e.elapsed), 0, "event", e))
    timeline.sort(key=lambda x: (x[0], x[1]))

    for _, _, kind, payload in timeline:
        if kind == "sample":
            s = payload
            row = [s.timestamp, plant, building, asset, sensor_label,
                   "Connected", s.elapsed,
                   _blank(s.rssi), _blank(s.rsrp), _blank(s.rsrq),
                   _blank(s.sinr), _blank(s.speed),
                   _blank(s.mccmnc), _blank(s.lac), _blank(s.cellid),
                   _blank(s.access_technology), _blank(s.band),
                   _blank(s.channel_type), _blank(s.sm_2g),
                   s.grade]
        else:
            e = payload
            row = [e.timestamp, plant, building, asset, sensor_label,
                   e.kind, e.elapsed] + [""] * (N_COLS - 8) + ["--"]
        ws.append(row)

    # Alternating row shading on data rows
    light = PatternFill("solid", fgColor="F0F4F8")
    white = PatternFill("solid", fgColor="FFFFFF")
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        fill = light if i % 2 == 0 else white
        for cell in row:
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = _CENTER

    # ── Summary block (appears at the BOTTOM, after all data rows) ──────
    ws.append([""] * N_COLS)

    # OVERALL RESULT — plain-English summary for a non-technical reader
    final_grade, final_speed = store.final_grade()
    n = store.count
    if n > 0 and final_grade != "--":
        headline = (f"OVERALL RESULT — {sensor_label}: {final_grade.upper()} signal "
                    f"(mean upload speed {final_speed} kB/s, top-3 of {n} samples)")
    else:
        headline = f"OVERALL RESULT — {sensor_label}: no clean samples collected"
    ws.append([headline] + [""] * (N_COLS - 2) + [final_grade])

    # FINAL SIGNAL QUALITY (top-3 mean)
    fg, fs = store.final_grade()
    ws.append([f"FINAL SIGNAL QUALITY  (mean of top-3 Upload Speeds"
               f"{': ' + str(fs) + ' kB/s' if fs else ''})"]
              + [""] * (N_COLS - 2) + [fg])

    # OVERALL SIGNAL QUALITY (all-sample mean)
    og, os_ = store.overall_grade()
    ws.append([f"OVERALL SIGNAL QUALITY  (mean of all Upload Speeds"
               f"{': ' + str(os_) + ' kB/s' if os_ else ''})"]
              + [""] * (N_COLS - 2) + [og])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    ws.row_dimensions[1].height = 22


def build_report_bytes(stores: Dict[int, SensorStore],
                       num_sensors: int,
                       plant: str, building: str, asset: str) -> io.BytesIO:
    """Generate the Excel workbook and return as a BytesIO buffer
    ready to send to the browser as a file download."""
    wb = Workbook()
    for name in list(wb.sheetnames):
        del wb[name]

    sheet_names = (["Sensor 1", "Sensor 2"] if num_sensors == 2
                   else ["Sensor 1"])
    for i, sname in enumerate(sheet_names, start=1):
        ws = wb.create_sheet(sname)
        _write_sheet(ws, stores[i], f"Sensor {i}", plant, building, asset)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf